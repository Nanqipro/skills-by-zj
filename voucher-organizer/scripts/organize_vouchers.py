#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
单据抽凭整理器 / Voucher batch organizer.

把"按凭证编号命名的文件夹"批量转换为"按目标列(如金蝶凭证号)命名的干净文件夹"：
  1) 解压文件夹内全部 ZIP（含嵌套 ZIP），保留中文文件名；
  2) 删除原始非 ZIP 文件与 ZIP 本体，只留下解压出来的单据文件；
  3) 依据台账 Excel 的映射，把文件夹重命名为目标列的值。

三段式，务必按顺序使用（destructive 操作前先 preflight 看清楚）：
  preflight  只扫描、只报告，不改动任何文件——确认映射齐全、无重名/无误删风险。
  process    真正执行 解压→清理→重命名。
  report     生成"原文件夹 → 新名 → 文件清单"的审计 CSV。

示例:
  python organize_vouchers.py preflight --config config.json
  python organize_vouchers.py process   --config config.json
  python organize_vouchers.py report    --config config.json --out 处理结果对照表.csv

所有参数都可用命令行覆盖 config.json。中文列名直接用原样字符串。
"""
import argparse
import csv
import json
import os
import re
import sys
import zipfile


# --------------------------------------------------------------------------- #
# 映射构建
# --------------------------------------------------------------------------- #
def build_mapping(cfg):
    """读取台账，返回 {源文件夹名: 目标名} 。

    源文件夹名 = key_cols 各列的值用连字符拼接（与磁盘上的文件夹名一致）。
    目标名     = target_col 的值。
    空目标值不会覆盖已存在的真实值（台账常有同一凭证编号的多行、部分目标为空）。
    """
    import openpyxl
    wb = openpyxl.load_workbook(cfg["xlsx"], data_only=True)
    ws = wb[cfg["sheet"]] if cfg.get("sheet") else wb.active
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col_idx(name):
        if name not in header:
            sys.exit(f"❌ 台账中找不到列「{name}」，现有列: {header}")
        return header.index(name) + 1

    key_idx = [col_idx(c) for c in cfg["key_cols"]]
    tgt_idx = col_idx(cfg["target_col"])
    filters = [(col_idx(k), str(v)) for k, v in cfg.get("filters", {}).items()]

    def fmt(v):
        # Excel 数字常带 .0；凭证编号等按整数处理
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    mapping, dup_conflict = {}, []
    for r in range(2, ws.max_row + 1):
        if any(str(ws.cell(r, ci).value) != val for ci, val in filters):
            continue
        if any(ws.cell(r, ci).value in (None, "") for ci in key_idx):
            continue
        key = "-".join(fmt(ws.cell(r, ci).value) for ci in key_idx)
        tgt = ws.cell(r, tgt_idx).value
        tgt = "" if tgt is None else str(tgt).strip()
        if key in mapping and mapping[key] and tgt and mapping[key] != tgt:
            dup_conflict.append((key, mapping[key], tgt))
        if tgt or key not in mapping:
            mapping[key] = tgt
    return mapping, dup_conflict


# --------------------------------------------------------------------------- #
# ZIP 处理
# --------------------------------------------------------------------------- #
def real_name(zi):
    """正确解码 ZIP 成员名：UTF-8 标志位优先，否则按 GBK/UTF-8 兜底（解决中文乱码）。"""
    if zi.flag_bits & 0x800:
        return zi.filename
    raw = zi.filename.encode("cp437", errors="replace")
    for enc in ("gbk", "utf-8"):
        try:
            return raw.decode(enc)
        except Exception:
            pass
    return zi.filename


def _unique(folder, name):
    base, ext = os.path.splitext(name)
    cand, i = name, 1
    while os.path.exists(os.path.join(folder, cand)):
        cand, i = f"{base}__{i}{ext}", i + 1
    return cand


def extract_folder(folder, log):
    """递归解压 folder 内全部 ZIP（扁平化、保留中文名），返回解压出的文件名集合。

    同名不同内容的成员会保留两份（加 __N 后缀）以防数据丢失。ZIP 本体处理后即删除。
    """
    extracted, rounds = set(), 0
    while True:
        zips = [f for f in os.listdir(folder) if f.lower().endswith(".zip")]
        if not zips:
            break
        rounds += 1
        if rounds > 20:
            log.append(("ERROR", folder, "嵌套层数异常，停止"))
            break
        for z in zips:
            zp = os.path.join(folder, z)
            try:
                with zipfile.ZipFile(zp) as zf:
                    for zi in zf.infolist():
                        if zi.is_dir():
                            continue
                        name = os.path.basename(real_name(zi))
                        if not name:
                            continue
                        data = zf.read(zi)
                        dest = os.path.join(folder, name)
                        if os.path.exists(dest) and os.path.getsize(dest) != len(data):
                            name = _unique(folder, name)
                            dest = os.path.join(folder, name)
                            log.append(("COLLISION", folder, name))
                        with open(dest, "wb") as fo:
                            fo.write(data)
                        extracted.add(name)
                os.remove(zp)  # 删除已解压的 ZIP 本体
            except zipfile.BadZipFile:
                log.append(("BAD-ZIP", folder, z))
    return extracted


# --------------------------------------------------------------------------- #
# 扫描 / 处理
# --------------------------------------------------------------------------- #
def list_source_folders(cfg):
    rx = re.compile(cfg["folder_regex"])
    return sorted(
        d for d in os.listdir(cfg["root"])
        if os.path.isdir(os.path.join(cfg["root"], d)) and rx.match(d)
    )


def preflight(cfg):
    mapping, dup_conflict = build_mapping(cfg)
    folders = list_source_folders(cfg)
    existing = set(os.listdir(cfg["root"]))
    illegal = re.compile(r'[\\/:*?"<>|]')

    not_in_map, empty_tgt, bad_name, tgt_exists, no_zip = [], [], [], [], []
    multizip = 0
    seen_targets = {}
    target_dups = []
    for f in folders:
        if f not in mapping:
            not_in_map.append(f)
            continue
        tgt = mapping[f]
        if not tgt:
            empty_tgt.append(f)
            continue
        if illegal.search(tgt):
            bad_name.append((f, tgt))
        if tgt in seen_targets:
            target_dups.append((seen_targets[tgt], f, tgt))
        seen_targets[tgt] = f
        if tgt in existing and tgt != f:
            tgt_exists.append((f, tgt))
        p = os.path.join(cfg["root"], f)
        zips = [x for x in os.listdir(p) if x.lower().endswith(".zip")]
        if not zips:
            no_zip.append(f)
        if len(zips) > 1:
            multizip += 1

    print("==================== 预检报告 (preflight) ====================")
    print(f"台账映射条数            : {len(mapping)}")
    print(f"待处理源文件夹          : {len(folders)}")
    print(f"  ├ 可正常处理          : {len(folders) - len(not_in_map) - len(empty_tgt)}")
    print(f"  ├ 含多个ZIP           : {multizip}")
    print(f"  └ 无ZIP(将保留原文件) : {len(no_zip)}  {no_zip if no_zip else ''}")
    print("--------------------- 风险检查 (应全为'无') ---------------------")
    print(f"台账中找不到映射        : {not_in_map or '无'}")
    print(f"目标列为空(将跳过)      : {empty_tgt or '无'}")
    print(f"目标名含非法字符        : {bad_name or '无'}")
    print(f"目标名彼此冲突(会覆盖)  : {target_dups or '无'}")
    print(f"目标名已存在            : {tgt_exists or '无'}")
    print(f"映射键冲突(同键不同值)  : {dup_conflict or '无'}")
    blocking = bad_name or target_dups or tgt_exists or dup_conflict
    print("============================================================")
    if blocking:
        print("⚠️  存在阻断性风险，请先人工核对台账后再 process。")
    elif not_in_map or empty_tgt:
        print("ℹ️  有源文件夹缺映射/目标为空，process 时会自动跳过这些文件夹。")
    else:
        print("✅ 预检通过，可以安全执行 process。")
    return not blocking


def process(cfg):
    mapping, _ = build_mapping(cfg)
    folders = list_source_folders(cfg)
    log, ok, skipped, kept = [], 0, 0, 0
    for f in folders:
        src = os.path.join(cfg["root"], f)
        tgt = mapping.get(f)
        if not tgt:
            log.append(("SKIP", f, "无映射或目标为空"))
            skipped += 1
            continue
        dst = os.path.join(cfg["root"], tgt)
        if os.path.exists(dst) and os.path.abspath(dst) != os.path.abspath(src):
            log.append(("SKIP", f, f"目标已存在 {tgt}"))
            skipped += 1
            continue

        before = set(os.listdir(src))
        has_zip = any(x.lower().endswith(".zip") for x in before)
        if has_zip:
            extracted = extract_folder(src, log)
            # 删除"原有的、且未被解压重新生成的"非ZIP文件（封面/凭证等散落文件）
            for item in before:
                if item.lower().endswith(".zip") or item in extracted:
                    continue
                fp = os.path.join(src, item)
                if os.path.isfile(fp):
                    os.remove(fp)
            log.append(("OK", f"{f} -> {tgt}", f"解压得到{len(extracted)}个文件"))
            ok += 1
        else:
            # 无ZIP：不要清空文件夹，仅去掉 .DS_Store，保留原文件
            ds = os.path.join(src, ".DS_Store")
            if os.path.exists(ds):
                os.remove(ds)
            log.append(("KEEP", f"{f} -> {tgt}", "无ZIP，保留原有文件"))
            kept += 1
        # 收尾：去掉 .DS_Store 再重命名
        ds = os.path.join(src, ".DS_Store")
        if os.path.exists(ds):
            os.remove(ds)
        os.rename(src, dst)

    for row in log:
        print(" | ".join(str(x) for x in row))
    print(f"\n处理完成：成功解压 {ok}，无ZIP保留 {kept}，跳过 {skipped}，合计 {len(folders)}。")
    if any(r[0] in ("SKIP", "BAD-ZIP", "ERROR", "COLLISION") for r in log):
        print("⚠️  存在需关注项（SKIP/BAD-ZIP/COLLISION），请查看上方日志。")


def report(cfg, out):
    mapping, _ = build_mapping(cfg)
    # 反查：目标名 -> 源文件夹名
    rev = {v: k for k, v in mapping.items() if v}
    present = set(os.listdir(cfg["root"]))
    rows = []
    for tgt, src in sorted(rev.items()):
        fp = os.path.join(cfg["root"], tgt)
        if not os.path.isdir(fp):
            continue
        files = sorted(x for x in os.listdir(fp) if not x.startswith("."))
        # 仍保留"散落原始单据"(如 单据凭证…-1.pdf / 申请单封面…-4.pdf) 说明该文件夹原本无ZIP
        looks_original = re.compile(r"(单据凭证|申请单封面).*-\d+\.pdf$", re.I)
        note = "无ZIP,保留原始单据" if any(looks_original.search(x) for x in files) else "已解压"
        rows.append([src, tgt, len(files), "; ".join(files), note])
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["原文件夹", "新文件夹名(目标列)", "文件数量", "文件清单", "备注"])
        w.writerows(rows)
    print(f"审计报告已生成：{out}（{len(rows)} 条记录）")


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def load_cfg(args):
    cfg = {
        "root": None,
        "xlsx": None,
        "sheet": None,
        "key_cols": ["会计年度", "凭证类型", "凭证编号"],
        "target_col": "金蝶凭证号",
        "filters": {},
        "folder_regex": r"^\d{4}-[A-Za-z0-9]+-\d+$",
    }
    if args.config:
        with open(args.config, encoding="utf-8") as f:
            cfg.update(json.load(f))
    for k in ("root", "xlsx", "sheet", "target_col", "folder_regex"):
        if getattr(args, k, None):
            cfg[k] = getattr(args, k)
    if args.key_cols:
        cfg["key_cols"] = [c.strip() for c in args.key_cols.split(",")]
    if args.filter:
        for item in args.filter:
            k, v = item.split("=", 1)
            cfg["filters"][k] = v
    missing = [k for k in ("root", "xlsx") if not cfg.get(k)]
    if missing:
        sys.exit(f"❌ 缺少必填参数: {missing}（用 --config 或命令行指定）")
    return cfg


def main():
    p = argparse.ArgumentParser(description="单据抽凭整理器：解压+清理+按台账重命名")
    p.add_argument("command", choices=["preflight", "process", "report"])
    p.add_argument("--config", help="config.json 路径")
    p.add_argument("--root", help="存放各凭证子文件夹的目录")
    p.add_argument("--xlsx", help="台账 Excel 路径")
    p.add_argument("--sheet", help="工作表名")
    p.add_argument("--key-cols", dest="key_cols", help="拼成文件夹名的列，逗号分隔（顺序敏感）")
    p.add_argument("--target-col", dest="target_col", help="作为新文件夹名的列")
    p.add_argument("--filter", action="append", help="行过滤 列名=值，可重复，如 公司代码=0111")
    p.add_argument("--folder-regex", dest="folder_regex", help="识别源文件夹的正则")
    p.add_argument("--out", default="处理结果对照表.csv", help="report 输出 CSV 路径")
    args = p.parse_args()

    cfg = load_cfg(args)
    if args.command == "preflight":
        ok = preflight(cfg)
        sys.exit(0 if ok else 1)
    elif args.command == "process":
        process(cfg)
    elif args.command == "report":
        report(cfg, args.out)


if __name__ == "__main__":
    main()
